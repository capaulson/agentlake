"""Excel (.xlsx, .xls) file adapter using openpyxl."""

from __future__ import annotations

from pathlib import Path

from agentlake.adapters.base import ExtractedContent, StructureMarker, TextBlock


class XlsxAdapter:
    """Extracts text from Excel spreadsheets.

    Each sheet is processed row-by-row.  Source locators follow the pattern
    ``sheet:{name}:row:{n}`` (1-indexed).
    """

    supported_extensions: list[str] = [".xlsx", ".xls"]
    supported_mimetypes: list[str] = [
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "application/vnd.ms-excel",
    ]

    def can_handle(self, filename: str, content_type: str) -> bool:
        return (
            Path(filename).suffix.lower() in self.supported_extensions
            or content_type in self.supported_mimetypes
        )

    def extract(self, file_bytes: bytes, filename: str) -> ExtractedContent:
        """Extract rows from each sheet as text blocks.

        Args:
            file_bytes: Raw Excel bytes.
            filename: Original filename.

        Returns:
            Extracted content with per-row text blocks.
        """
        import io

        import openpyxl

        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)

        metadata: dict = {
            "filename": filename,
            "sheet_names": wb.sheetnames,
            "sheet_count": len(wb.sheetnames),
        }

        blocks: list[TextBlock] = []
        structure: list[StructureMarker] = []
        position = 0

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            structure.append(
                StructureMarker(
                    marker_type="section_start",
                    position=position,
                    metadata={"sheet": sheet_name},
                )
            )

            for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
                cell_values = [
                    str(cell) if cell is not None else "" for cell in row
                ]
                # Skip entirely empty rows
                if not any(v.strip() for v in cell_values):
                    continue

                row_text = " | ".join(cell_values)
                blocks.append(
                    TextBlock(
                        content=row_text,
                        block_type="table",
                        position=position,
                        source_locator=f"sheet:{sheet_name}:row:{row_idx}",
                        metadata={"sheet": sheet_name, "row": row_idx},
                    )
                )
                position += 1

        wb.close()

        return ExtractedContent(
            text_blocks=blocks,
            metadata=metadata,
            structure=structure,
        )
